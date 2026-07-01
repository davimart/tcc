# ============================================================
# EXPORTAR NOTEBOOK COMO .PY
# ============================================================
# Esta célula exporta automaticamente todas as células de
# código do notebook para um arquivo .py reutilizável.
# ============================================================

import json
import pandas as pd

ipynb_path = "func_aux_corrigido.ipynb"
py_path = "func_aux.py"

with open(ipynb_path, "r", encoding="utf-8") as f:
    notebook = json.load(f)

code_cells = []

for cell in notebook["cells"]:

    if cell["cell_type"] == "code":

        code = "".join(cell["source"])

        code_cells.append(code)
        code_cells.append("\n\n")

with open(py_path, "w", encoding="utf-8") as f:
    f.writelines(code_cells)

print(f"Arquivo exportado: {py_path}")

import numpy as np


def taxa_real_por_grupo(
    y_true,
    sensitive_features,
    grupo_privilegiado,
    grupo_desprivilegiado,
    label_favoravel=1
):
    """
    Calcula a taxa real do label favorável em cada grupo sensível.

    AR|P=1 = P(Y = label_favoravel | A = grupo_privilegiado)
    AR|P=0 = P(Y = label_favoravel | A = grupo_desprivilegiado)
    """

    y_true = np.asarray(y_true)
    sensitive_features = np.asarray(sensitive_features)

    mask_priv = sensitive_features == grupo_privilegiado
    mask_despriv = sensitive_features == grupo_desprivilegiado

    actual_rate_privilegiado = np.mean(y_true[mask_priv] == label_favoravel) if mask_priv.sum() > 0 else np.nan
    actual_rate_desprivilegiado = np.mean(y_true[mask_despriv] == label_favoravel) if mask_despriv.sum() > 0 else np.nan

    return {
        "actual_rate_privilegiado": actual_rate_privilegiado,
        "actual_rate_desprivilegiado": actual_rate_desprivilegiado
    }



def avaliar_tradicional(modelo, X_test, y_test, atributo_sensivel=None):
    from sklearn.metrics import accuracy_score, precision_score, recall_score, f1_score
    from fairlearn.postprocessing import ThresholdOptimizer

    if isinstance(modelo, ThresholdOptimizer):
        y_pred = modelo.predict(
            X_test,
            sensitive_features=X_test[atributo_sensivel]
        )
    else:
        y_pred = modelo.predict(X_test)

    return {
        "accuracy": accuracy_score(y_test, y_pred),
        "precision": precision_score(y_test, y_pred, zero_division=0),
        "recall": recall_score(y_test, y_pred, zero_division=0),
        "f1": f1_score(y_test, y_pred, zero_division=0),
    }

from fairlearn.postprocessing import ThresholdOptimizer

def avaliar_fairness(
    modelo,
    X_test,
    y_test,
    atributo_sensivel,
    grupo_privilegiado,
    grupo_desprivilegiado
):
    import numpy as np
    from sklearn.metrics import confusion_matrix

    if isinstance(modelo, ThresholdOptimizer):
        y_pred = modelo.predict(
            X_test,
            sensitive_features=X_test[atributo_sensivel]
        )
    else:
        y_pred = modelo.predict(X_test)

    grupo = X_test[atributo_sensivel]

    metricas_grupo = {}

    for nome, valor in {
        "privilegiado": grupo_privilegiado,
        "desprivilegiado": grupo_desprivilegiado
    }.items():

        idx = grupo == valor

        y_true_g = y_test[idx]
        y_pred_g = y_pred[idx]

        tn, fp, fn, tp = confusion_matrix(
            y_true_g,
            y_pred_g,
            labels=[0, 1]
        ).ravel()

        tpr = tp / (tp + fn) if (tp + fn) > 0 else np.nan
        fpr = fp / (fp + tn) if (fp + tn) > 0 else np.nan
        positive_rate = (tp + fp) / len(y_pred_g) if len(y_pred_g) > 0 else np.nan

        metricas_grupo[nome] = {
            "tpr": tpr,
            "fpr": fpr,
            "positive_rate": positive_rate
        }

    tpr_priv = metricas_grupo["privilegiado"]["tpr"]
    tpr_despriv = metricas_grupo["desprivilegiado"]["tpr"]

    fpr_priv = metricas_grupo["privilegiado"]["fpr"]
    fpr_despriv = metricas_grupo["desprivilegiado"]["fpr"]

    pr_priv = metricas_grupo["privilegiado"]["positive_rate"]
    pr_despriv = metricas_grupo["desprivilegiado"]["positive_rate"]

    return {
        "equal_opportunity": abs(tpr_priv - tpr_despriv),
        "equalized_odds": 0.5 * (
            abs(tpr_priv - tpr_despriv)
            + abs(fpr_priv - fpr_despriv)
        ),
        "demographic_parity": abs(pr_priv - pr_despriv),

        "tpr_privilegiado": tpr_priv,
        "tpr_desprivilegiado": tpr_despriv,
        "fpr_privilegiado": fpr_priv,
        "fpr_desprivilegiado": fpr_despriv,
        "positive_rate_privilegiado": pr_priv,
        "positive_rate_desprivilegiado": pr_despriv,
    }

def avaliar_modelo(
    modelo,
    X_test,
    y_test,
    atributo_sensivel,
    grupo_privilegiado,
    grupo_desprivilegiado
):
    metricas = {}

    metricas.update(
        avaliar_tradicional(
            modelo,
            X_test,
            y_test,
            atributo_sensivel=atributo_sensivel
        )
    )

    metricas.update(
        avaliar_fairness(
            modelo,
            X_test,
            y_test,
            atributo_sensivel,
            grupo_privilegiado,
            grupo_desprivilegiado
        )
    )

    return metricas

def detectar_tipo_tarefa(y):
    import numpy as np
    y_array = np.array(y)

    if np.issubdtype(y_array.dtype, np.floating):
        return "regressao"

    n_classes = len(np.unique(y_array))
    if n_classes == 2:
        return "binaria"
    elif n_classes > 2:
        return "multiclasse"
    return "desconhecido"


def desempenho_tradicional_binario(y_true, y_pred, explain=False):
    from sklearn.metrics import accuracy_score, precision_score, recall_score, f1_score
    
    acc = accuracy_score(y_true, y_pred)
    print(f"Acurácia:  {acc:.4f}")       

    prec = precision_score(y_true, y_pred)
    print(f"Precisão:  {prec:.4f}")

    rec = recall_score(y_true, y_pred)
    print(f"Recall:    {rec:.4f}")
    
    f1 = f1_score(y_true, y_pred)
    print(f"F1-Score:  {f1:.4f}")

    if explain:
        print("🔹 Acurácia: proporção de acertos no total de exemplos. Boa se > 0.80, mas pode enganar em dados desbalanceados.")
        print("🔹 Precisão: entre os que foram preditos como positivos, quantos realmente são. Boa se você quer evitar falsos positivos.")
        print("🔹 Recall: entre os positivos reais, quantos foram detectados. Boa se você quer evitar falsos negativos.") 
        print("🔹 F1-Score: média harmônica entre precisão e recall. Equilibra ambos quando há desbalanceamento.")
    
    resultado = pd.DataFrame([{
    "Acurácia": acc,
    "Precisão": prec,
    "Recall": rec,
    "F1-Score": f1
    }])

    return resultado

def desempenho_tradicional_multiclasse(y_true, y_pred, explain=False):
    from sklearn.metrics import accuracy_score, precision_score, recall_score, f1_score
    
    acc = accuracy_score(y_true, y_pred)
    print(f"Acurácia:  {acc:.4f}")
    if explain:
        print("🔹 Acurácia: proporção de acertos entre todas as classes. Boa se > 0.80, mas cuidado com desbalanceamento.")

    prec = precision_score(y_true, y_pred, average='macro')
    print(f"Precisão (macro): {prec:.4f}")
    if explain:
        print("🔹 Precisão: média da precisão de cada classe. Mostra se o modelo é justo com todas as classes.")

    rec = recall_score(y_true, y_pred, average='macro')
    print(f"Recall (macro):   {rec:.4f}")
    if explain:
        print("🔹 Recall: média do recall de cada classe. Mede cobertura média de cada classe verdadeira.")

    f1 = f1_score(y_true, y_pred, average='macro')
    print(f"F1-Score (macro): {f1:.4f}")
    if explain:
        print("🔹 F1-Score: média harmônica entre precisão e recall para todas as classes.")
    return acc, prec, rec, f1

def desempenho_tradicional_regressao(y_true, y_pred, explain=False):
    from sklearn.metrics import mean_squared_error, mean_absolute_error, r2_score

    r2 = r2_score(y_true, y_pred)
    print(f"R2:  {r2:.4f}")
    if explain:
        print("🔹 R2: quão bem o modelo explica a variância. 1.0 é perfeito. Pode ser negativo se o modelo for ruim.")

    mse = mean_squared_error(y_true, y_pred)
    print(f"MSE: {mse:.4f}")
    if explain:
        print("🔹 MSE: erro quadrático médio. Penaliza mais os erros grandes. Quanto menor, melhor.")

    mae = mean_absolute_error(y_true, y_pred)
    print(f"MAE: {mae:.4f}")
    if explain:
        print("🔹 MAE: erro absoluto médio. Indica o erro médio em termos absolutos. Quanto menor, melhor.")
    return r2, mse, mae

def desempenho_tradicional(model, X_test, y_test, explain=False):
    tipo = detectar_tipo_tarefa(y_test)

    y_pred = model.predict(X_test)

    if tipo == "binaria":
        return desempenho_tradicional_binario(y_test, y_pred, explain)
    elif tipo == "multiclasse":
        return desempenho_tradicional_multiclasse(y_test, y_pred, explain)
    elif tipo == "regressao":
        return desempenho_tradicional_regressao(y_test, y_pred, explain)
    else:
        print("⚠️ Tipo de tarefa não reconhecido.")
        return None


def summarize_dataset(df, nome, txt=True, save_images=True):
    """
    📊 Gera um sumário completo de estatísticas e estrutura do dataset.

    Parâmetros:
    - df: pandas DataFrame
    - nome: nome identificador do dataset (ex: 'cluster_0')
    - txt: se True, salva o sumário em sumario/{nome}.txt
    - save_images: se True, salva gráficos em sumario/{nome}/
    """
    import pandas as pd
    import numpy as np
    import matplotlib.pyplot as plt
    import seaborn as sns
    import os
    import sys

    # ⚙️ Limite para considerar colunas categóricas
    MAX_UNIQUE_VALUES_FOR_CAT = 3

    # Diretórios e caminhos
    base_dir = "sumario"
    os.makedirs(base_dir, exist_ok=True)

    txt_path = os.path.join(base_dir, f"{nome}.txt") if txt else None
    image_dir = os.path.join(base_dir, nome) if save_images else None

    if image_dir:
        os.makedirs(image_dir, exist_ok=True)

    f = open(txt_path, "w", encoding="utf-8") if txt_path else None

    def p_save(text):
        print(text)
        if f: f.write(text + "\n")

    p_save(f"📘 Sumário para: {nome}")
    p_save("=" * 40)

    # 📦 Tamanho
    p_save("\n📦 Tamanho do dataset:")
    p_save(f"- Linhas: {df.shape[0]}")
    p_save(f"- Colunas: {df.shape[1]}")

    # 📄 Primeiras linhas
    print("\n📄 Primeiras 5 linhas:")
    display(df.head())

    # 🔢 Informações básicas
    print("\n🔢 Informações básicas:")
    df.info(buf=sys.stdout if not f else f)

    # 📈 Estatísticas descritivas
    p_save("\n📈 Estatísticas descritivas:")
    desc = df.describe().T
    display(desc)
    if f: f.write(desc.to_string(float_format="%.2f") + "\n")

    # 🧾 Colunas categóricas
    cat_cols = [
        col for col in df.columns
        if df[col].nunique() <= MAX_UNIQUE_VALUES_FOR_CAT
        and col != 'target'
        and df[col].dtype in [np.int64, np.int32, np.bool_, object]
    ]
    if cat_cols:
        print("\n🧾 Frequência das colunas categóricas:")
        cat_summary = pd.DataFrame()
        for col in cat_cols:
            freqs = df[col].value_counts(dropna=False).to_frame(name=col)
            cat_summary = pd.concat([cat_summary, freqs], axis=1)
        display(cat_summary)

    # 📊 Correlação
    p_save("\n📊 Correlação entre variáveis numéricas:")
    corr = df.corr(numeric_only=True)
    display(corr.style.background_gradient(cmap='coolwarm', axis=None).format(precision=2))
    if f: f.write("\n" + corr.to_string(float_format="%.2f") + "\n")

    # 🎨 Gráficos gerais
    if save_images:
        num_cols = df.select_dtypes(include=np.number).columns
        if len(num_cols) > 0:
            # Boxplot
            plt.figure(figsize=(12, 6))
            sns.boxplot(data=df[num_cols])
            plt.xticks(rotation=45)
            plt.title(f"Boxplot - {nome}")
            plt.tight_layout()
            plt.savefig(os.path.join(image_dir, f"{nome}_boxplot.png"))
            plt.close()

            # Histogramas
            for col in num_cols:
                plt.figure(figsize=(6, 4))
                sns.histplot(df[col], kde=True, bins=20, color='skyblue')
                plt.title(f"{col} - {nome}")
                plt.tight_layout()
                plt.savefig(os.path.join(image_dir, f"{nome}_{col}_hist.png"))
                plt.close()

    # 🎯 Análise da variável de destino
    target_cols = [col for col in df.columns if 'target' in col.lower()]
    for target_col in target_cols:
        print(f"\n🎯 Análise da variável de destino: '{target_col}'")

        # Distribuição dos valores
        distrib = df[target_col].value_counts(dropna=False)
        print("\nDistribuição dos valores:")
        display(distrib.to_frame(name='Frequência'))

        # Correlação com variáveis numéricas
        if pd.api.types.is_numeric_dtype(df[target_col]):
            corrs = df.corr(numeric_only=True)[target_col].drop(target_col)
            top_corrs = corrs.abs().sort_values(ascending=False).head(5)
            print(f"\n🔍 Top 5 correlações com '{target_col}':")
            display(corrs.loc[top_corrs.index].to_frame(name='Correlação'))

            for col in top_corrs.index:
                print(f"\n📊 Média de '{col}' por categoria de '{target_col}':")
                grouped = df.groupby(target_col)[col].mean().to_frame(name='Média')
                display(grouped)

                if save_images:
                    plt.figure(figsize=(6, 4))
                    sns.histplot(df[col], kde=True, bins=20, hue=df[target_col], palette='Set1')
                    plt.title(f"{col} por {target_col} - {nome}")
                    plt.tight_layout()
                    plt.savefig(os.path.join(image_dir, f"{nome}_{col}_por_{target_col}_hist.png"))
                    plt.close()

    if f:
        f.close()
        print(f"\n📁 Resumo salvo em: {txt_path}")



def consistency_score(X, y_pred, k=5):
    from sklearn.metrics import pairwise_distances
    from sklearn.neighbors import NearestNeighbors
    import numpy as np
    """
    📏 Mede Individual Fairness com base na Consistência.
    Quanto maior, mais justo o modelo.
    
    Parâmetros:
    - X: Features (sem atributo sensível)
    - y_pred: Saídas do modelo (classes ou probabilidades)
    - k: Número de vizinhos

    Retorna:
    - Média da similaridade de predições entre vizinhos
    """
    nn = NearestNeighbors(n_neighbors=k+1)  # +1 porque inclui o próprio ponto
    nn.fit(X)
    neighbors = nn.kneighbors(X, return_distance=False)[:, 1:]  # Remove o próprio índice

    diffs = []
    for i in range(X.shape[0]):
        neighbor_preds = y_pred[neighbors[i]]
        diffs.append(np.mean(np.abs(y_pred[i] - neighbor_preds)))

    return 1 - np.mean(diffs)


def desempenho_fairness_binario(y_true, y_pred, grupo_sensivel, explain=False):
    """
    Calcula métricas de fairness (EO, EOdds, DP) para classificação binária.
    
    Parâmetros:
    - y_true: Series com rótulos verdadeiros
    - y_pred: lista ou array com predições do modelo
    - grupo_sensivel: Series (uma coluna) ou DataFrame (várias colunas sensíveis binárias)
    - explain: se True, imprime os resultados
    
    Retorna:
    - DataFrame com métricas de fairness para cada coluna sensível
    """
    import numpy as np
    import pandas as pd
    from sklearn.metrics import confusion_matrix

    # Garante que y_pred tem mesmo índice de y_true
    y_pred = pd.Series(y_pred, index=y_true.index)

    # Se grupo_sensivel for uma única coluna, converte para DataFrame
    if isinstance(grupo_sensivel, pd.Series):
        grupo_sensivel = grupo_sensivel.to_frame()

    resultados = {}

    for coluna in grupo_sensivel.columns:
        grupos = grupo_sensivel[coluna].unique()
        if len(grupos) <= 1:
            continue  # ignora coluna constante

        tprs, fprs, positives = [], [], []

        for valor in grupos:
            idx = grupo_sensivel[coluna] == valor
            y_true_g = y_true[idx]
            y_pred_g = y_pred[idx]

            tn, fp, fn, tp = confusion_matrix(y_true_g, y_pred_g, labels=[0, 1]).ravel()

            tpr = tp / (tp + fn) if (tp + fn) > 0 else np.nan
            fpr = fp / (fp + tn) if (fp + tn) > 0 else np.nan
            dp  = (tp + fp) / len(y_pred_g) if len(y_pred_g) > 0 else np.nan

            tprs.append(tpr)
            fprs.append(fpr)
            positives.append(dp)

        resultados[coluna] = {
            "Equal Opportunity": np.nanmax(tprs) - np.nanmin(tprs),
            "Equalized Odds": ((np.nanmax(tprs) - np.nanmin(fprs)) +
                               (np.nanmax(fprs) - np.nanmin(fprs))) / 2,
            "Demographic Parity": np.nanmax(positives) - np.nanmin(positives)
        }

        if explain:
            print(f"\n🔍 Grupo sensível: {coluna}")
            print(f"📏 Equal Opportunity:     {resultados[coluna]['Equal Opportunity']:.4f}")
            print(f"📏 Equalized Odds:        {resultados[coluna]['Equalized Odds']:.4f}")
            print(f"📏 Demographic Parity:    {resultados[coluna]['Demographic Parity']:.4f}")

    return pd.DataFrame(resultados).T


import json
import pandas as pd
import numpy as np

from pathlib import Path
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


def exportar_metricas_monografia_xlsx(
    df_metricas,
    caminho_saida,
    caminho_dicionario,
    casas_decimais=3
):
    """
    Exporta a tabela de métricas para .xlsx com formatação visual real.

    A função:
    - ordena os modelos;
    - renomeia colunas;
    - renomeia valores categóricos;
    - formata números;
    - aplica negrito nos melhores valores das métricas principais;
    - ignora nulo e base na escolha dos melhores valores.

    Parâmetros
    ----------
    df_metricas : pd.DataFrame
        Tabela bruta de métricas.

    caminho_saida : str ou Path
        Caminho do arquivo .xlsx de saída.

    caminho_dicionario : str ou Path
        Caminho do JSON de conversão.

    casas_decimais : int
        Número de casas decimais na tabela exportada.
    """

    caminho_saida = Path(caminho_saida)

    with open(caminho_dicionario, "r", encoding="utf-8") as f:
        dic = json.load(f)

    df = df_metricas.copy()


    # ========================================================
    # Ordenação
    # ========================================================

    ordem_modelos = dic.get("ordem_modelos", {})

    if "modelo" in df.columns:
        df["_ordem_modelo"] = df["modelo"].map(ordem_modelos)

    colunas_ordenacao = []

    if "dataset" in df.columns:
        colunas_ordenacao.append("dataset")

    if "_ordem_modelo" in df.columns:
        colunas_ordenacao.append("_ordem_modelo")

    if "config_sensivel" in df.columns:
        colunas_ordenacao.append("config_sensivel")

    if colunas_ordenacao:
        df = df.sort_values(by=colunas_ordenacao)

    if "_ordem_modelo" in df.columns:
        df = df.drop(columns=["_ordem_modelo"])


    # ========================================================
    # Guarda versão original para decidir negritos
    # ========================================================

    df_original = df.copy()


    # ========================================================
    # Renomeia valores categóricos
    # ========================================================

    if "datasets" in dic and "dataset" in df.columns:
        df["dataset"] = df["dataset"].replace(dic["datasets"])

    if "modelos" in dic and "modelo" in df.columns:
        df["modelo"] = df["modelo"].replace(dic["modelos"])

    if "mitigacoes" in dic and "mitigacao" in df.columns:
        df["mitigacao"] = df["mitigacao"].replace(dic["mitigacoes"])

    if "configs_sensiveis" in dic and "config_sensivel" in df.columns:
        df["config_sensivel"] = df["config_sensivel"].replace(
            dic["configs_sensiveis"]
        )


    # ========================================================
    # Ordena/seleciona colunas se estiverem no dicionário
    # ========================================================

    colunas_dict = dic.get("colunas", {})

    colunas_existentes_ordenadas = [
        col for col in colunas_dict.keys()
        if col in df.columns
    ]

    colunas_restantes = [
        col for col in df.columns
        if col not in colunas_existentes_ordenadas
    ]

    df = df[colunas_existentes_ordenadas + colunas_restantes]


    # ========================================================
    # Renomeia colunas para exibição
    # ========================================================

    mapa_colunas_existente = {
        col_original: col_novo
        for col_original, col_novo in colunas_dict.items()
        if col_original in df.columns
    }

    df = df.rename(columns=mapa_colunas_existente)


    # ========================================================
    # Exporta para Excel
    # ========================================================

    with pd.ExcelWriter(caminho_saida, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="metricas")

        wb = writer.book
        ws = writer.sheets["metricas"]


        # ====================================================
        # Estilos básicos
        # ====================================================

        header_fill = PatternFill(
            fill_type="solid",
            fgColor="D9EAF7"
        )

        thin_gray = Side(style="thin", color="D9D9D9")

        border = Border(
            left=thin_gray,
            right=thin_gray,
            top=thin_gray,
            bottom=thin_gray
        )

        header_font = Font(bold=True)
        normal_font = Font(bold=False)
        bold_font = Font(bold=True)

        center = Alignment(horizontal="center", vertical="center")
        left = Alignment(horizontal="left", vertical="center")


        # Cabeçalho
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = border

        # Corpo
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.border = border
                cell.alignment = center


        # ====================================================
        # Formatação numérica
        # ====================================================

        formato_numero = "0." + ("0" * casas_decimais)

        nomes_colunas_excel = [
            cell.value for cell in ws[1]
        ]

        for col_idx, nome_coluna in enumerate(nomes_colunas_excel, start=1):
            for row_idx in range(2, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)

                if isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool):
                    cell.number_format = formato_numero


        # ====================================================
        # Aplica negrito nos melhores valores
        # ====================================================

        metricas_maior_melhor = dic.get("metricas_maior_melhor", [])
        metricas_menor_melhor = dic.get("metricas_menor_melhor", [])
        linhas_ignorar_negrito = dic.get("linhas_ignorar_negrito", [])

        metricas_para_negrito = metricas_maior_melhor + metricas_menor_melhor

        # Mapeia nome original -> nome exibido no Excel
        mapa_nome_exibido = {
            col_original: colunas_dict.get(col_original, col_original)
            for col_original in df_original.columns
        }

        # Grupos de comparação:
        # Como nulo e base aparecem só uma vez, e as mitigações podem aparecer
        # por config, agrupamos por dataset.
        # Assim o melhor FairShap/FAGTB/Threshold é escolhido dentro do dataset.
        grupos = ["dataset"] if "dataset" in df_original.columns else []

        if grupos:
            iter_grupos = df_original.groupby(grupos, dropna=False)
        else:
            iter_grupos = [(None, df_original)]

        for _, grupo_df in iter_grupos:

            grupo_comparar = grupo_df[
                ~grupo_df["modelo"].isin(linhas_ignorar_negrito)
            ]

            if grupo_comparar.empty:
                continue

            for metrica in metricas_para_negrito:

                if metrica not in grupo_comparar.columns:
                    continue

                valores = pd.to_numeric(
                    grupo_comparar[metrica],
                    errors="coerce"
                )

                if valores.dropna().empty:
                    continue

                if metrica in metricas_maior_melhor:
                    melhor_valor = valores.max()
                    indices_melhores = valores[
                        valores == melhor_valor
                    ].index

                elif metrica in metricas_menor_melhor:
                    melhor_valor = valores.abs().min()
                    indices_melhores = valores[
                        valores.abs() == melhor_valor
                    ].index

                else:
                    continue

                nome_coluna_excel = mapa_nome_exibido.get(metrica, metrica)

                if nome_coluna_excel not in nomes_colunas_excel:
                    continue

                col_excel = nomes_colunas_excel.index(nome_coluna_excel) + 1

                for idx in indices_melhores:
                    # +2 porque Excel começa em 1 e linha 1 é cabeçalho
                    row_excel = df_original.index.get_loc(idx) + 2
                    ws.cell(row=row_excel, column=col_excel).font = bold_font


        # ====================================================
        # Ajuste de largura de colunas
        # ====================================================

        for col_idx, col_cells in enumerate(ws.columns, start=1):
            max_len = 0

            for cell in col_cells:
                value = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(value))

            width = min(max(max_len + 2, 10), 24)
            ws.column_dimensions[get_column_letter(col_idx)].width = width


        # Congela cabeçalho
        ws.freeze_panes = "A2"

        # Filtro
        ws.auto_filter.ref = ws.dimensions

    return caminho_saida

def preparar_X_numerico_para_shap(X):
    """
    Garante que X seja um DataFrame numérico para explicadores caixa-preta.
    """
    import pandas as pd

    X_numeric = X.copy()

    for col in X_numeric.columns:
        X_numeric[col] = pd.to_numeric(
            X_numeric[col],
            errors="coerce"
        )

    X_numeric = X_numeric.fillna(0.0).astype(float)

    return X_numeric

def modelo_suportado_tree_explainer(model):
    """
    Verifica se o modelo pode usar SHAP TreeExplainer diretamente.
    """
    from sklearn.tree import DecisionTreeClassifier, DecisionTreeRegressor
    from sklearn.ensemble import GradientBoostingClassifier, GradientBoostingRegressor

    return isinstance(model, (
        GradientBoostingClassifier,
        GradientBoostingRegressor,
        DecisionTreeClassifier,
        DecisionTreeRegressor
    ))

def calcular_shap_modelo(
    model,
    X,
    background_size=100,
    explain_size=500,
    random_state=42
):
    """
    Calcula SHAP values para um modelo.

    Usa TreeExplainer quando possível.
    Caso contrário, usa explicabilidade model-agnostic.

    Retorna um dicionário com:
    - shap_values
    - X_plot
    - explainer
    - metodo
    """

    import shap
    import numpy as np
    import pandas as pd

    X_exp = X.copy()

    # ============================================================
    # Caminho rápido: TreeExplainer
    # ============================================================

    if modelo_suportado_tree_explainer(model):

        explainer = shap.TreeExplainer(model)
        shap_values = explainer.shap_values(X_exp)
        X_plot = X_exp

        return {
            "shap_values": shap_values,
            "X_plot": X_plot,
            "explainer": explainer,
            "metodo": "tree"
        }

    # ============================================================
    # Caminho genérico: modelo caixa-preta
    # Ex.: FAGTB, Threshold wrappers, outros modelos customizados
    # ============================================================

    print("⚠️ Tipo de modelo não suportado pelo TreeExplainer.")
    print("🔁 Usando explicabilidade model-agnostic (caixa-preta).")

    X_numeric = preparar_X_numerico_para_shap(X_exp)

    X_background = shap.sample(
        X_numeric,
        min(background_size, len(X_numeric)),
        random_state=random_state
    )

    X_plot = shap.sample(
        X_numeric,
        min(explain_size, len(X_numeric)),
        random_state=random_state
    )

    def predict_score(X_input):

        X_input = pd.DataFrame(
            X_input,
            columns=X_numeric.columns
        )

        if hasattr(model, "predict_proba"):
            return model.predict_proba(X_input)[:, 1]

        return np.asarray(model.predict(X_input))

    explainer = shap.Explainer(
        predict_score,
        X_background
    )

    shap_exp = explainer(X_plot)
    shap_values = shap_exp.values

    if shap_values.shape != X_plot.shape:
        raise AssertionError(
            f"shap_values.shape = {shap_values.shape}, "
            f"X_plot.shape = {X_plot.shape}"
        )

    return {
        "shap_values": shap_values,
        "X_plot": X_plot,
        "explainer": explainer,
        "metodo": "agnostic"
    }

def selecionar_shap_values_classe(
    shap_values,
    classe=1
):
    """
    Retorna uma matriz 2D de SHAP values.

    Casos tratados:
    - shap_values 2D: retorna direto.
    - shap_values list: retorna shap_values[classe].
    - shap_values 3D: retorna shap_values[:, :, classe].

    Por padrão usa classe=1, que normalmente representa
    a classe favorável/positiva.
    """

    import numpy as np

    if isinstance(shap_values, list):
        return shap_values[classe]

    shap_values = np.asarray(shap_values)

    if shap_values.ndim == 2:
        return shap_values

    if shap_values.ndim == 3:
        return shap_values[:, :, classe]

    raise ValueError(
        f"Formato de shap_values não reconhecido: {shap_values.shape}"
    )

def calcular_importancia_media_shap(
    shap_values,
    nomes_features,
    classe=1
):
    """
    Calcula mean(|SHAP|) por variável.
    """

    import numpy as np
    import pandas as pd

    shap_2d = selecionar_shap_values_classe(
        shap_values,
        classe=classe
    )

    mean_abs = np.abs(shap_2d).mean(axis=0)

    return pd.DataFrame({
        "variavel": nomes_features,
        "mean_abs_shap": mean_abs
    }).sort_values(
        "mean_abs_shap",
        ascending=False
    )

def plot_shap_bar(
    shap_values,
    X_plot,
    max_display=15,
    titulo=None,
    classe=1
):
    """
    Gera bar plot SHAP.
    """

    import shap
    import matplotlib.pyplot as plt

    shap_2d = selecionar_shap_values_classe(
        shap_values,
        classe=classe
    )

    plt.figure()

    shap.summary_plot(
        shap_2d,
        X_plot,
        plot_type="bar",
        max_display=max_display,
        show=False
    )

    if titulo:
        plt.title(titulo)

    return plt.gcf()

def plot_shap_beeswarm(
    shap_values,
    X_plot,
    max_display=15,
    titulo=None,
    classe=1,
    cmap="coolwarm"
):
    """
    Gera beeswarm plot SHAP.
    """

    import shap
    import matplotlib.pyplot as plt

    shap_2d = selecionar_shap_values_classe(
        shap_values,
        classe=classe
    )

    plt.figure()

    shap.summary_plot(
        shap_2d,
        X_plot,
        max_display=max_display,
        cmap=plt.get_cmap(cmap),
        show=False
    )

    if titulo:
        plt.title(titulo)

    return plt.gcf()

def explain_model(
    model,
    X,
    max_display=15,
    nome=None,
    target_name=None,
    cmap="coolwarm",
    classe=1,
    explicar_todas_classes=False,
    background_size=100,
    explain_size=500
):
    """
    Calcula e plota explicações SHAP globais.

    Retorna:
    - figuras
    - explicacao_shap

    explicacao_shap contém:
    - shap_values
    - X_plot
    - explainer
    - metodo
    """

    import numpy as np

    figuras = {}

    explicacao = calcular_shap_modelo(
        model=model,
        X=X,
        background_size=background_size,
        explain_size=explain_size
    )

    shap_values = explicacao["shap_values"]
    X_plot = explicacao["X_plot"]


    # ============================================================
    # Decide se precisa explicar várias classes
    # ============================================================

    n_classes = None

    if isinstance(shap_values, list):
        n_classes = len(shap_values)

    else:
        shap_array = np.asarray(shap_values)

        if shap_array.ndim == 3:
            n_classes = shap_array.shape[2]


    # ============================================================
    # Explica todas as classes, se solicitado
    # ============================================================

    if explicar_todas_classes and n_classes is not None:

        for i in range(n_classes):

            nome_classe = (
                f"{target_name}: {i}"
                if target_name is not None
                else f"Classe {i}"
            )

            titulo_bar = (
                f"{nome} — {nome_classe} — Importância global"
                if nome
                else f"{nome_classe} — Importância global"
            )

            titulo_beeswarm = (
                f"{nome} — {nome_classe} — Distribuição dos impactos"
                if nome
                else f"{nome_classe} — Distribuição dos impactos"
            )

            figuras[f"classe_{i}_bar"] = plot_shap_bar(
                shap_values=shap_values,
                X_plot=X_plot,
                max_display=max_display,
                titulo=titulo_bar,
                classe=i
            )

            figuras[f"classe_{i}_beeswarm"] = plot_shap_beeswarm(
                shap_values=shap_values,
                X_plot=X_plot,
                max_display=max_display,
                titulo=titulo_beeswarm,
                classe=i,
                cmap=cmap
            )

        return figuras, explicacao


    # ============================================================
    # Explica apenas uma classe
    # ============================================================

    titulo_bar = (
        f"{nome} — Importância global"
        if nome
        else "Importância global"
    )

    titulo_beeswarm = (
        f"{nome} — Distribuição dos impactos"
        if nome
        else "Distribuição dos impactos"
    )

    figuras["bar"] = plot_shap_bar(
        shap_values=shap_values,
        X_plot=X_plot,
        max_display=max_display,
        titulo=titulo_bar,
        classe=classe
    )

    figuras["beeswarm"] = plot_shap_beeswarm(
        shap_values=shap_values,
        X_plot=X_plot,
        max_display=max_display,
        titulo=titulo_beeswarm,
        classe=classe,
        cmap=cmap
    )

    return figuras, explicacao

def plot_shap_barras_agrupadas(
    explicacoes_por_modelo,
    top_por_modelo=7,
    top_final=10,
    classe=1,
    titulo="Comparação da importância relativa dos atributos via SHAP",
    salvar_em=None,
    salvar_tabela_ranking_em=None,
    salvar_tabela_plot_em=None,
    figsize=(7, 12),
    normalizar=True,
    ordem_modelos=None,
    incluir_base=True,
    mostrar_mensagem=True
):
    """
    Gera gráfico de barras agrupadas horizontais comparando importâncias SHAP.

    A função:
    - calcula mean(|SHAP|) por modelo;
    - opcionalmente normaliza a importância dentro de cada modelo;
    - seleciona variáveis pela união dos top N de cada modelo;
    - limita a top_final variáveis;
    - gera gráfico horizontal, mais adequado para página em orientação retrato;
    - gera uma tabela de ranking no formato:
        Posição | Modelo nome | Modelo % | ...

    Parâmetros
    ----------
    explicacoes_por_modelo : dict
        Dicionário no formato:
        {
            "Base": exp_base,
            "FairShap": exp_fairshap,
            "FAGTB": exp_fagtb,
            "Threshold": exp_threshold
        }

    top_por_modelo : int
        Número de variáveis mais importantes consideradas por modelo.

    top_final : int
        Número máximo de variáveis exibidas no gráfico final.

    classe : int
        Classe SHAP explicada.

    titulo : str
        Título do gráfico.

    salvar_em : str ou Path
        Caminho para salvar o gráfico.

    salvar_tabela_ranking_em : str ou Path
        Caminho para salvar a tabela de ranking.

    salvar_tabela_plot_em : str ou Path
        Caminho para salvar a matriz usada no gráfico.

    figsize : tuple
        Tamanho da figura. Para monografia, algo como (7, 12) costuma funcionar bem.

    normalizar : bool
        Se True, usa importância relativa:
            mean(|SHAP|) / soma(mean(|SHAP|))

    ordem_modelos : list
        Ordem desejada dos modelos.

    incluir_base : bool
        Se False, remove Base mesmo que esteja em explicacoes_por_modelo.

    mostrar_mensagem : bool
        Se True, imprime diagnóstico sobre top_por_modelo e top_final.

    Retorna
    -------
    df_shap : pd.DataFrame
        Tabela longa com importância por variável e modelo.

    tabela_plot : pd.DataFrame
        Tabela pivotada usada no gráfico.

    tabela_ranking_formatada : pd.DataFrame
        Tabela no formato adequado para monografia.
    """

    import pandas as pd
    import matplotlib.pyplot as plt

    if ordem_modelos is None:
        ordem_modelos = ["Base", "FairShap", "FAGTB", "Threshold"]

    if not incluir_base:
        ordem_modelos = [m for m in ordem_modelos if m != "Base"]

    registros = []

    # ============================================================
    # Calcula importância média por modelo
    # ============================================================

    for nome_modelo, exp in explicacoes_por_modelo.items():

        if nome_modelo not in ordem_modelos:
            continue

        shap_values = exp["shap_values"]
        X_plot = exp["X_plot"]

        df_imp = calcular_importancia_media_shap(
            shap_values=shap_values,
            nomes_features=X_plot.columns.tolist(),
            classe=classe
        )

        soma = df_imp["mean_abs_shap"].sum()

        if normalizar and soma != 0:
            df_imp["importancia"] = df_imp["mean_abs_shap"] / soma
        else:
            df_imp["importancia"] = df_imp["mean_abs_shap"]

        df_imp["modelo"] = nome_modelo

        df_imp["ranking"] = (
            df_imp["importancia"]
            .rank(method="min", ascending=False)
            .astype(int)
        )

        registros.append(df_imp)

    df_shap = pd.concat(registros, ignore_index=True)

    modelos_existentes = [
        m for m in ordem_modelos
        if m in df_shap["modelo"].unique()
    ]

    df_shap = df_shap[df_shap["modelo"].isin(modelos_existentes)].copy()

    df_shap["modelo"] = pd.Categorical(
        df_shap["modelo"],
        categories=modelos_existentes,
        ordered=True
    )

    # ============================================================
    # Seleciona top N de cada modelo
    # ============================================================

    variaveis_top_por_modelo = {}

    for nome_modelo in modelos_existentes:

        top_vars = (
            df_shap[df_shap["modelo"] == nome_modelo]
            .sort_values("importancia", ascending=False)
            .head(top_por_modelo)["variavel"]
            .tolist()
        )

        variaveis_top_por_modelo[nome_modelo] = top_vars

    variaveis_top_uniao = set()

    for vars_modelo in variaveis_top_por_modelo.values():
        variaveis_top_uniao.update(vars_modelo)

    n_variaveis_uniao = len(variaveis_top_uniao)

    # ============================================================
    # Se passar do limite, mantém top global relativo
    # ============================================================

    df_global = (
        df_shap[df_shap["variavel"].isin(variaveis_top_uniao)]
        .groupby("variavel", as_index=False)["importancia"]
        .mean()
        .sort_values("importancia", ascending=False)
    )

    variaveis_finais = (
        df_global
        .head(top_final)["variavel"]
        .tolist()
    )

    variaveis_removidas = sorted(
        list(variaveis_top_uniao - set(variaveis_finais))
    )

    if mostrar_mensagem:
        print("=" * 80)
        print("Seleção de variáveis para gráfico SHAP agrupado")
        print(f"Modelos considerados: {modelos_existentes}")
        print(f"Top por modelo: {top_por_modelo}")
        print(f"Variáveis únicas na união dos tops: {n_variaveis_uniao}")
        print(f"Limite final de variáveis no gráfico: {top_final}")

        if n_variaveis_uniao <= top_final:
            print("✅ Todas as variáveis do top de cada modelo foram mantidas.")
        else:
            print(
                f"⚠️ Nem todas as variáveis do top {top_por_modelo} "
                f"de cada modelo couberam no top_final={top_final}."
            )
            print(f"Variáveis removidas: {variaveis_removidas}")

        print("=" * 80)

    # ============================================================
    # Monta tabela para plot
    # ============================================================

    df_plot = df_shap[
        df_shap["variavel"].isin(variaveis_finais)
    ].copy()

    tabela_plot = df_plot.pivot_table(
        index="variavel",
        columns="modelo",
        values="importancia",
        fill_value=0
    )

    tabela_plot = tabela_plot[modelos_existentes]

    ordem_variaveis = (
        tabela_plot.mean(axis=1)
        .sort_values(ascending=False)
        .index
        .tolist()
    )

    tabela_plot = tabela_plot.loc[ordem_variaveis]

    # ============================================================
    # Tabela de ranking no formato desejado
    # ============================================================

    linhas_ranking = []

    max_posicao = top_final

    for posicao in range(1, max_posicao + 1):

        linha = {
            "Posição": posicao
        }

        for modelo in modelos_existentes:

            df_modelo = (
                df_shap[df_shap["modelo"] == modelo]
                .sort_values("importancia", ascending=False)
                .reset_index(drop=True)
            )

            if posicao <= len(df_modelo):

                variavel = df_modelo.loc[posicao - 1, "variavel"]
                importancia = df_modelo.loc[posicao - 1, "importancia"]

                linha[f"{modelo} nome"] = variavel
                linha[f"{modelo} %"] = importancia * 100

            else:

                linha[f"{modelo} nome"] = ""
                linha[f"{modelo} %"] = ""

        linhas_ranking.append(linha)

    tabela_ranking_formatada = pd.DataFrame(linhas_ranking)

    if salvar_tabela_ranking_em is not None:
        tabela_ranking_formatada.to_csv(
            salvar_tabela_ranking_em,
            index=False
        )

    if salvar_tabela_plot_em is not None:
        tabela_plot.to_csv(salvar_tabela_plot_em)

    # ============================================================
    # Gráfico horizontal verticalizado
    # ============================================================

    # Inverte a ordem para a variável mais importante aparecer no topo
# ============================================================
# Gráfico horizontal verticalizado
# ============================================================

    tabela_plot_plotar = tabela_plot.iloc[::-1]

    cores_modelos = {
        "Base": "#9E9E9E",
        "FairShap": "#6BAED6",
        "FAGTB": "#3182BD",
        "Threshold": "#08519C"
    }

    cores_usadas = [
        cores_modelos.get(modelo, None)
        for modelo in tabela_plot_plotar.columns
    ]

    ax = tabela_plot_plotar.plot(
        kind="barh",
        figsize=figsize,
        width=0.82,
        color=cores_usadas
    )

    ax.set_title(titulo)
    ax.set_xlabel(
        "Importância relativa SHAP"
        if normalizar
        else "mean(|SHAP|)"
    )
    ax.set_ylabel("Variável")


    # ============================================================
    # Adiciona ranking nas barras
    # ============================================================

    for container, modelo in zip(ax.containers, tabela_plot_plotar.columns):

        for bar, variavel in zip(container, tabela_plot_plotar.index):

            valor = bar.get_width()

            if valor <= 0:
                continue

            ranking = df_shap[
                (df_shap["modelo"] == modelo) &
                (df_shap["variavel"] == variavel)
            ]["ranking"]

            if ranking.empty:
                continue

            ranking = int(ranking.iloc[0])

            ax.text(
                valor + 0.005,
                bar.get_y() + bar.get_height() / 2,
                f"#{ranking}",
                va="center",
                ha="left",
                fontsize=8
            )


    # ============================================================
    # Espaço extra para os rankings
    # ============================================================

    xmax = tabela_plot_plotar.max().max()
    ax.set_xlim(0, xmax * 1.15)


    # ============================================================
    # Legenda ordenada
    # ============================================================

    handles, labels = ax.get_legend_handles_labels()

    ordem_legenda = ["Base", "FairShap", "FAGTB", "Threshold"]

    ordem_indices = [
        labels.index(nome)
        for nome in ordem_legenda
        if nome in labels
    ]

    handles_ordenados = [handles[i] for i in ordem_indices]
    labels_ordenados = [labels[i] for i in ordem_indices]

    ax.legend(
        handles_ordenados,
        labels_ordenados,
        title="Modelo",
        loc="lower right"
    )

    plt.tight_layout()

    if salvar_em is not None:
        plt.savefig(
            salvar_em,
            dpi=300,
            bbox_inches="tight"
        )

    plt.show()

    return df_shap, tabela_plot, tabela_ranking_formatada

def explain_model_velho(
    model,
    X,
    max_display=15,
    nome=None,
    target_name=None,
    cmap="coolwarm"
):
    import shap
    import numpy as np
    import matplotlib.pyplot as plt
    import pandas as pd

    from sklearn.tree import DecisionTreeClassifier, DecisionTreeRegressor
    from sklearn.ensemble import GradientBoostingClassifier, GradientBoostingRegressor

    # ============================================================
    # Configuração interna para modelos caixa-preta
    # ============================================================

    background_size = 100
    explain_size = 500

    X_exp = X.copy()

    # ============================================================
    # Modelos suportados diretamente pelo TreeExplainer
    # ============================================================

    modelos_problema = isinstance(model, DecisionTreeClassifier)

    modelos_sem_problema = isinstance(model, (
        GradientBoostingClassifier,
        DecisionTreeRegressor,
        GradientBoostingRegressor
    ))

    modelos_tree_explainer = isinstance(model, (
        GradientBoostingClassifier,
        DecisionTreeRegressor,
        GradientBoostingRegressor,
        DecisionTreeClassifier
    ))

    figuras = {}

    # ============================================================
    # Caminho rápido: TreeExplainer
    # ============================================================

    if modelos_tree_explainer:

        explainer = shap.TreeExplainer(model)
        shap_values = explainer.shap_values(X_exp)
        X_plot = X_exp

    # ============================================================
    # Caminho genérico: modelo caixa-preta
    # Ex.: FAGTB
    # ============================================================

    else:

        print("⚠️ Tipo de modelo não suportado pelo TreeExplainer.")
        print("🔁 Usando explicabilidade model-agnostic (caixa-preta).")

        # ============================================================
        # Garante dados puramente numéricos para o SHAP caixa-preta
        # ============================================================

        X_numeric = X_exp.copy()

        for col in X_numeric.columns:
            X_numeric[col] = pd.to_numeric(
                X_numeric[col],
                errors="coerce"
            )

        X_numeric = X_numeric.fillna(0.0).astype(float)

        X_background = shap.sample(
            X_numeric,
            min(background_size, len(X_numeric)),
            random_state=42
        )

        X_plot = shap.sample(
            X_numeric,
            min(explain_size, len(X_numeric)),
            random_state=42
        )

        def predict_score(X_input):

            # O SHAP às vezes passa numpy array.
            # Convertemos de volta para DataFrame com os nomes originais.
            X_input = pd.DataFrame(
                X_input,
                columns=X_numeric.columns
            )

            if hasattr(model, "predict_proba"):
                return model.predict_proba(X_input)[:, 1]

            return np.asarray(model.predict(X_input))

        explainer = shap.Explainer(
            predict_score,
            X_background
        )

        shap_exp = explainer(X_plot)
        shap_values = shap_exp.values

        if shap_values.shape != X_plot.shape:
            raise AssertionError(
                f"shap_values.shape = {shap_values.shape}, "
                f"X_plot.shape = {X_plot.shape}"
            )

        modelos_problema = False
        modelos_sem_problema = True

    # ============================================================
    # Caso padrão: modelos sem problema
    # ============================================================

    if not modelos_problema:

        if not modelos_sem_problema:
            print("⚠️ Tipo de modelo não identificado, tentando procedimento padrão")

            if shap_values.shape != X_plot.shape:
                raise AssertionError(
                    f"shap_values.shape = {shap_values.shape}, "
                    f"X.shape = {X_plot.shape}"
                )

        print("🔍 Gerando explicabilidade global do modelo...")

        plt.figure()

        shap.summary_plot(
            shap_values,
            X_plot,
            plot_type="bar",
            max_display=max_display,
            show=False
        )

        if nome:
            plt.title(f"{nome} — Importância global")

        figuras["bar"] = plt.gcf()

        plt.figure()

        shap.summary_plot(
            shap_values,
            X_plot,
            max_display=max_display,
            cmap=plt.get_cmap(cmap),
            show=False
        )

        if nome:
            plt.title(f"{nome} — Distribuição dos impactos")

        figuras["beeswarm"] = plt.gcf()

        return figuras

    # ============================================================
    # Caso especial: DecisionTreeClassifier
    # ============================================================

    elif isinstance(model, DecisionTreeClassifier):

        if isinstance(shap_values, list):

            n_classes = len(shap_values)

            for i in range(n_classes):

                print(f"\n📊 {target_name}: {i} — Explicação global")

                plt.figure()

                shap.summary_plot(
                    shap_values[i],
                    X_plot,
                    plot_type="bar",
                    max_display=max_display,
                    show=False
                )

                if nome:
                    plt.title(
                        f"{nome} — {target_name}: {i} — Importância global"
                    )

                figuras[f"classe_{i}_bar"] = plt.gcf()

                print(f"📈 {target_name}: {i} — Distribuição dos impactos")

                plt.figure()

                shap.summary_plot(
                    shap_values[i],
                    X_plot,
                    max_display=max_display,
                    cmap=plt.get_cmap(cmap),
                    show=False
                )

                if nome:
                    plt.title(
                        f"{nome} — {target_name}: {i} — Distribuição dos impactos"
                    )

                figuras[f"classe_{i}_beeswarm"] = plt.gcf()

            return figuras

        else:

            _, _, n_classes = shap_values.shape

            for i in range(n_classes):

                print(f"\n📊 {target_name}: {i} — Explicação global")

                plt.figure()

                shap.summary_plot(
                    shap_values[:, :, i],
                    X_plot,
                    plot_type="bar",
                    max_display=max_display,
                    show=False
                )

                if nome:
                    plt.title(
                        f"{nome} — {target_name}: {i} — Importância global"
                    )

                figuras[f"classe_{i}_bar"] = plt.gcf()

                print(f"📈 {target_name}: {i} — Distribuição dos impactos")

                plt.figure()

                shap.summary_plot(
                    shap_values[:, :, i],
                    X_plot,
                    max_display=max_display,
                    cmap=plt.get_cmap(cmap),
                    show=False
                )

                if nome:
                    plt.title(
                        f"{nome} — {target_name}: {i} — Distribuição dos impactos"
                    )

                figuras[f"classe_{i}_beeswarm"] = plt.gcf()

            return figuras

def explain_individual(index, model, X_train):
    import shap
    from scipy.special import expit
    import matplotlib.pyplot as plt
    """
    🧠 Explicabilidade local com SHAP (versão waterfall).
    Mostra o impacto de cada feature + exibe a probabilidade prevista no gráfico.
    """
    # Obter valores SHAP
    explainer = shap.TreeExplainer(model)
    shap_values = explainer.shap_values(X_train)
    instance = X_train.iloc[[index]]

    # Corrigir acesso ao escalar
    expected_value = explainer.expected_value[0]
    fx = float(shap_values[index].sum() + expected_value)
    prob = float(expit(fx))

    # Construir explicação
    explanation = shap.Explanation(
        values=shap_values[index],
        base_values=expected_value,
        data=instance.values[0],
        feature_names=instance.columns
    )

    # Plot com título seguro (sem emoji)
    shap.plots.waterfall(explanation, show=False)
    plt.title(f"SHAP Waterfall — Instância {index} | Prob. prevista: {prob:.2f}", fontsize=12)
    plt.show()


def explicar_pesos_fairshap(
    X_train,
    y_train,
    pesos_fairshap,
    atributo_sensivel,
    nome_dataset,
    nome_config,
    max_depth_arvore=None,
    max_depth_plot=3,
    max_display=15,
    cmap="coolwarm"
):
    import numpy as np
    import pandas as pd
    import matplotlib.pyplot as plt
    import shap

    from sklearn.tree import DecisionTreeRegressor, plot_tree
    from sklearn.metrics import r2_score, mean_squared_error, mean_absolute_error

    X_train = X_train.reset_index(drop=True)
    y_train = y_train.reset_index(drop=True)
    pesos_fairshap = pd.Series(pesos_fairshap).reset_index(drop=True)

    df_aux = X_train.copy()
    df_aux["target"] = y_train
    df_aux["peso_fairshap"] = pesos_fairshap

    figuras = {}

    # 1. Distribuição global dos pesos
    fig, ax = plt.subplots(figsize=(7, 4))
    ax.hist(df_aux["peso_fairshap"], bins=30)
    ax.set_title(f"{nome_dataset} — Distribuição dos pesos FairShap")
    ax.set_xlabel("Peso FairShap")
    ax.set_ylabel("Frequência")
    figuras["distribuicao_pesos"] = fig

    # 2. Boxplot por grupo sensível e target
    fig, ax = plt.subplots(figsize=(9, 4))
    df_aux.boxplot(
        column="peso_fairshap",
        by=[atributo_sensivel, "target"],
        ax=ax
    )
    ax.set_title(f"{nome_dataset} — Pesos por grupo sensível e target")
    ax.set_xlabel(f"{atributo_sensivel}, target")
    ax.set_ylabel("Peso FairShap")
    fig.suptitle("")
    figuras["boxplot_sensivel_target"] = fig

    # 3. Árvore de decisão/regressão predizendo pesos
    arvore = DecisionTreeRegressor(
        max_depth=max_depth_arvore,
        random_state=42
    )

    arvore.fit(X_train, pesos_fairshap)

    pesos_pred = arvore.predict(X_train)

    metricas_arvore = {
        "dataset": nome_dataset,
        "config_sensivel": nome_config,
        "atributo_sensivel": atributo_sensivel,
        "r2": r2_score(pesos_fairshap, pesos_pred),
        "mse": mean_squared_error(pesos_fairshap, pesos_pred),
        "mae": mean_absolute_error(pesos_fairshap, pesos_pred),
        "max_depth_arvore": max_depth_arvore,
        "max_depth_plot": max_depth_plot,
        "n_amostras": len(X_train)
    }

    fig, ax = plt.subplots(figsize=(22, 10))
    plot_tree(
        arvore,
        feature_names=X_train.columns,
        filled=True,
        rounded=True,
        fontsize=8,
        max_depth=max_depth_plot,
        ax=ax
    )
    ax.set_title(
        f"{nome_dataset} — Recorte da árvore aproximando pesos FairShap"
    )
    figuras["arvore_pesos"] = fig

    # 4. SHAP da árvore: importância global
    explainer = shap.TreeExplainer(arvore)
    shap_values = explainer.shap_values(X_train)

    plt.figure()
    shap.summary_plot(
        shap_values,
        X_train,
        plot_type="bar",
        show=False,
        max_display=max_display
    )
    plt.title(f"{nome_dataset} — SHAP da árvore dos pesos")
    figuras["shap_arvore_bar"] = plt.gcf()

    # 5. SHAP da árvore: beeswarm
    plt.figure()
    shap.summary_plot(
        shap_values,
        X_train,
        show=False,
        max_display=max_display,
        cmap=plt.get_cmap(cmap)
    )
    plt.title(f"{nome_dataset} — Beeswarm SHAP da árvore dos pesos")
    figuras["shap_arvore_beeswarm"] = plt.gcf()

    return {
        "figuras": figuras,
        "arvore": arvore,
        "metricas_arvore": metricas_arvore
    }

def modelo_nulo(
    X_train,
    X_test,
    y_train,
    y_test,
    label="Modelo nulo",
    save_path=None
):
    import joblib

    from sklearn.dummy import DummyClassifier
    from sklearn.metrics import classification_report

    model = DummyClassifier(strategy="most_frequent")
    model.fit(X_train, y_train)

    y_pred = model.predict(X_test)

    print(f"\n🔍 Resultados para: {label}")
    print(classification_report(y_test, y_pred))

    if save_path is not None:
        joblib.dump(model, save_path)

    return model

def modelo_base(
    X_train,
    X_test,
    y_train,
    y_test,
    weights=None,
    label="Base",
    save_path=None
):
    """
    Treina um GradientBoostingClassifier.

    Parameters
    ----------
    X_train, X_test
    y_train, y_test
    weights : sample weights (ex.: FairShap)
    label : nome exibido no relatório
    save_path : caminho .joblib para salvar o modelo
                se None, não salva

    Returns
    -------
    model : modelo treinado
    """

    import joblib

    from sklearn.metrics import classification_report
    from sklearn.ensemble import GradientBoostingClassifier

    model = GradientBoostingClassifier(
        random_state=42
    )

    model.fit(
        X_train,
        y_train,
        sample_weight=weights
    )

    y_pred = model.predict(X_test)

    print(f"\n🔍 Resultados para: {label}")
    print(classification_report(y_test, y_pred))

    if save_path is not None:
        joblib.dump(model, save_path)

    return model

def pesos_fairshap(
    X_train,
    X_test,
    y_train,
    y_test,
    protected_attribute_col,
    privileged_value,
    unprivileged_value,
    label_favorable,
    label_unfavorable
):
    import numpy as np
    from fairSV.fair_shapley import FairShapley

    """
    Calcula os pesos do FairShap.

    Retorna:
    - weights: vetor de pesos normalizados
    - fair_sv_extractor: objeto FairShapley com os resultados
    """

    protected_values = X_test[protected_attribute_col].values

    protected_attributes_dict = {
        "values": protected_values,
        "privileged_protected_attribute": privileged_value,
        "unprivileged_protected_attribute": unprivileged_value,
        "favorable_label": label_favorable,
        "unfavorable_label": label_unfavorable
    }

    fair_sv_extractor = FairShapley(
        X_train.values,
        y_train,
        X_test.values,
        y_test,
        protected_attributes_dict=protected_attributes_dict,
        show_plot=False,
        calculate_2dim=True
    )

    best_k, _, _ = fair_sv_extractor.get_best_K()

    fair_sv_extractor.get_SV_matrix(K=best_k)
    fair_sv_extractor.get_sv_arrays()

    
    fair_sv_extractor.get_sv_arrays()

    # Equalized Odds no código do FairShap
    sv_fairness = fair_sv_extractor.sv_average_odds_difference

    min_sv = np.min(sv_fairness)
    max_sv = np.max(sv_fairness)

    if max_sv == min_sv:
        pesos = np.ones_like(sv_fairness, dtype=float)
    else:
        pesos = (sv_fairness - min_sv) / (max_sv - min_sv)

    return pesos, fair_sv_extractor

def modelo_fagtb(
    X_train,
    y_train,
    sensitive_train,
    X_test=None,
    y_test=None,
    sensitive_test=None,
    n_estimators=200,
    learning_rate=0.01,
    max_depth=3,
    max_features=None,
    lambda_fagtb=0.05
):
    import importlib
    import fagtb_core
    importlib.reload(fagtb_core)
    from fagtb_core import FAGTB

    modelo = FAGTB(
        n_estimators=n_estimators,
        learning_rate=learning_rate,
        min_samples_split=2,
        min_impurity=0,
        max_depth=max_depth,
        max_features=max_features,
        regression=False
    )

    modelo.fit(
        X_train,
        y_train,
        sensitive_train,
        LAMBDA=lambda_fagtb,
        Xtest=X_test if X_test is not None else X_train,
        yt=y_test if y_test is not None else y_train,
        sensitivet=sensitive_test if sensitive_test is not None else sensitive_train
    )

    return modelo

import numpy as np


class ThresholdWrapper:
    """
    Wrapper para fazer o ThresholdOptimizer funcionar como modelo comum.

    Ele guarda o atributo sensível usado no pós-processamento
    e permite chamar apenas:

        modelo.predict(X)

    em vez de:

        modelo.predict(X, sensitive_features=X[atributo_sensivel])
    """

    def __init__(self, modelo_threshold, atributo_sensivel):
        self.modelo_threshold = modelo_threshold
        self.atributo_sensivel = atributo_sensivel

    def predict(self, X):
        return self.modelo_threshold.predict(
            X,
            sensitive_features=X[self.atributo_sensivel]
        )

    def predict_proba(self, X):
        """
        O ThresholdOptimizer normalmente não produz probabilidade contínua real.
        Então aqui criamos uma saída compatível com sklearn baseada na decisão final.

        Retorno:
            coluna 0 = probabilidade/classe de 0
            coluna 1 = probabilidade/classe de 1
        """
        y_pred = self.predict(X)
        y_pred = np.asarray(y_pred)

        return np.column_stack([
            1 - y_pred,
            y_pred
        ])

    def get_threshold_model(self):
        return self.modelo_threshold

    def get_sensitive_attribute(self):
        return self.atributo_sensivel

